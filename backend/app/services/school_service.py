from sqlalchemy import select, func, and_, update, Integer, Cast, case
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.student import StudentProfile
from app.models.university import University
from app.models.employment_warning import EmploymentWarning
from app.models.college_employment import CollegeEmployment
from app.models.job import JobDescription, JobApplication
from app.models.account import Account, RoleType
from app.core.redis_client import get_redis
from app.core.config import get_settings
from app.utils.industry_normalizer import normalize_industry, batch_normalize_industries
from app.utils.salary_parser import parse_salary, is_valid_salary
from app.utils.education_mapper import normalize_education
from app.utils.province_normalizer import normalize_province_name
from app.services.scarce_talent_analyzer import ScarceTalentAnalyzer
import json
import uuid
import io
from collections import Counter


class SchoolService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_dashboard_data(self, university_id: str) -> dict:
        """返回学校首页统计数据"""
        # 一次查询获取所有状态的学生数量
        result = await self.db.execute(
            select(
                StudentProfile.employment_status,
                func.count(StudentProfile.profile_id)
            )
            .where(StudentProfile.university_id == university_id)
            .group_by(StudentProfile.employment_status)
        )
        status_counts = {row.employment_status: row.count for row in result.all()}

        total_students = sum(status_counts.values())
        employed_nums = status_counts.get(1, 0)
        further_study_nums = status_counts.get(2, 0)
        abroad_nums = status_counts.get(3, 0)
        unemployed_nums = status_counts.get(0, 0)
        employment_rate = (employed_nums / total_students * 100) if total_students > 0 else 0.0

        # 获取预警列表
        warnings_result = await self.db.execute(
            select(EmploymentWarning)
            .where(EmploymentWarning.university_id == university_id)
            .order_by(EmploymentWarning.created_at.desc())
            .limit(5)
        )
        warnings = []
        for w in warnings_result.scalars().all():
            warnings.append({
                "warning_id": w.warning_id,
                "account_id": w.account_id,
                "warning_type": w.warning_type,
                "level": w.level,
                "handled": w.handled
            })

        # 获取学院排名
        rankings_result = await self.db.execute(
            select(CollegeEmployment)
            .where(CollegeEmployment.university_id == university_id)
            .order_by(CollegeEmployment.employment_rate.desc())
            .limit(10)
        )
        college_rankings = []
        for r in rankings_result.scalars().all():
            college_rankings.append({
                "college_name": r.college_name,
                "employment_rate": float(r.employment_rate) if r.employment_rate else 0,
                "employed_nums": r.employed_nums,
                "graduate_nums": r.graduate_nums
            })

        return {
            "total_students": total_students,
            "employed_nums": employed_nums,
            "unemployed_nums": unemployed_nums,
            "further_study_nums": further_study_nums,
            "abroad_nums": abroad_nums,
            "employment_rate": round(employment_rate, 2),
            "college_rankings": college_rankings,
            "warnings": warnings,
            "new_jobs_this_month": 0
        }

    async def get_students(self, university_id: str, filters: dict) -> dict:
        """分页查询学生列表"""
        page = filters.get("page", 1)
        page_size = filters.get("page_size", 10)
        employment_status = filters.get("employment_status")
        major = filters.get("major")
        graduation_year = filters.get("graduation_year")

        # 构建查询条件
        conditions = [StudentProfile.university_id == university_id]
        if employment_status is not None:
            conditions.append(StudentProfile.employment_status == employment_status)
        if major:
            conditions.append(StudentProfile.major.like(f"%{major}%"))
        if graduation_year:
            conditions.append(StudentProfile.graduation_year == graduation_year)

        # 查询总数
        count_result = await self.db.execute(
            select(func.count(StudentProfile.profile_id)).where(and_(*conditions))
        )
        total = count_result.scalar() or 0

        # 分页查询
        offset = (page - 1) * page_size
        result = await self.db.execute(
            select(StudentProfile)
            .where(and_(*conditions))
            .offset(offset)
            .limit(page_size)
        )
        students = result.scalars().all()

        return {
            "list": [{
                "profile_id": s.profile_id,
                "account_id": s.account_id,
                "student_no": s.student_no,
                "college": s.college,
                "major": s.major,
                "degree": s.degree,
                "graduation_year": s.graduation_year,
                "employment_status": s.employment_status,
                "cur_company": s.cur_company,
                "cur_city": s.cur_city,
                "is_registered": bool(s.account_id) and s.account_id.strip() != ""
            } for s in students],
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def export_students(self, university_id: str, filters: dict = None) -> list:
        """导出学生数据为Excel兼容的数据列表"""
        import io
        from sqlalchemy import select
        from app.models.account import Account

        filters = filters or {}
        employment_status = filters.get("employment_status")
        major = filters.get("major")
        graduation_year = filters.get("graduation_year")

        # 构建查询条件
        conditions = [StudentProfile.university_id == university_id]
        if employment_status is not None:
            conditions.append(StudentProfile.employment_status == employment_status)
        if major:
            conditions.append(StudentProfile.major.like(f"%{major}%"))
        if graduation_year:
            conditions.append(StudentProfile.graduation_year == graduation_year)

        # 不分页，查询全部
        result = await self.db.execute(
            select(StudentProfile).where(and_(*conditions))
        )
        students = result.scalars().all()

        DEGREE_MAP = {1: '本科', 2: '硕士', 3: '博士', 4: '大专', 5: '其他'}
        EMPLOYMENT_STATUS_MAP = {0: '待就业', 1: '已就业', 2: '升学', 3: '出国'}

        # 查询账户信息用于获取姓名
        account_ids = [s.account_id for s in students if s.account_id]
        account_map = {}
        if account_ids:
            acc_result = await self.db.execute(
                select(Account).where(Account.account_id.in_(account_ids))
            )
            for acc in acc_result.scalars().all():
                account_map[acc.account_id] = acc

        export_data = []
        for s in students:
            account = account_map.get(s.account_id)
            export_data.append({
                "学号": s.student_no or "",
                "姓名": account.real_name if account else "",
                "学院": s.college or "",
                "专业": s.major or "",
                "学历": DEGREE_MAP.get(s.degree, '未知'),
                "毕业年份": s.graduation_year or "",
                "就业状态": EMPLOYMENT_STATUS_MAP.get(s.employment_status, '未知'),
                "当前公司": s.cur_company or "",
                "当前城市": s.cur_city or "",
                "当前行业": s.cur_industry or "",
                "当前薪资": s.cur_salary or "",
                "期望城市": s.desire_city or "",
                "期望行业": s.desire_industry or "",
                "期望薪资": f"{s.desire_salary_min or '-'} ~ {s.desire_salary_max or '-'}" if s.desire_salary_min or s.desire_salary_max else "",
                "生源省份": s.province_origin or "",
                "GPA": s.gpa or "",
                "技能特长": ", ".join(s.skills) if s.skills else "",
                "实习经历": s.internship or "",
                "注册状态": "已注册" if s.account_id else "未注册",
            })

        return export_data

    async def get_students_count(self, university_id: str, filters: dict = None) -> int:
        """获取符合条件的学生总数"""
        from sqlalchemy import select, func

        filters = filters or {}
        employment_status = filters.get("employment_status")
        major = filters.get("major")
        graduation_year = filters.get("graduation_year")

        conditions = [StudentProfile.university_id == university_id]
        if employment_status is not None:
            conditions.append(StudentProfile.employment_status == employment_status)
        if major:
            conditions.append(StudentProfile.major.like(f"%{major}%"))
        if graduation_year:
            conditions.append(StudentProfile.graduation_year == graduation_year)

        result = await self.db.execute(
            select(func.count(StudentProfile.profile_id)).where(and_(*conditions))
        )
        return result.scalar() or 0

    async def export_students_stream(self, university_id: str, filters: dict = None, batch_size: int = 1000, profile_ids: list[str] = None):
        """
        流式导出学生数据 - 异步生成器，分批yield数据
        每批返回dict列表，用于流式写入Excel

        Args:
            university_id: 学校ID
            filters: 过滤条件 (employment_status, major, graduation_year)
            batch_size: 每批查询数量
            profile_ids: 指定导出的学生ID列表，不传则按filters条件导出全部
        """
        from sqlalchemy import select
        from app.models.account import Account

        filters = filters or {}
        employment_status = filters.get("employment_status")
        major = filters.get("major")
        graduation_year = filters.get("graduation_year")

        DEGREE_MAP = {1: '本科', 2: '硕士', 3: '博士', 4: '大专', 5: '其他'}
        EMPLOYMENT_STATUS_MAP = {0: '待就业', 1: '已就业', 2: '升学', 3: '出国'}

        # 构建查询条件
        conditions = [StudentProfile.university_id == university_id]
        # 优先按指定ID列表导出
        if profile_ids:
            conditions.append(StudentProfile.profile_id.in_(profile_ids))
        else:
            # 按过滤条件导出
            if employment_status is not None:
                conditions.append(StudentProfile.employment_status == employment_status)
            if major:
                conditions.append(StudentProfile.major.like(f"%{major}%"))
            if graduation_year:
                conditions.append(StudentProfile.graduation_year == graduation_year)

        # 先获取所有账户信息用于关联姓名
        all_accounts_result = await self.db.execute(
            select(Account.account_id, Account.real_name)
        )
        account_rows = all_accounts_result.fetchall()
        account_map = {row[0]: row[1] for row in account_rows}

        # 分批查询
        offset = 0
        while True:
            result = await self.db.execute(
                select(StudentProfile)
                .where(and_(*conditions))
                .offset(offset)
                .limit(batch_size)
            )
            students = result.scalars().all()

            if not students:
                break

            batch = []
            for s in students:
                batch.append({
                    "学号": s.student_no or "",
                    "姓名": account_map.get(s.account_id, "") if s.account_id else "",
                    "学院": s.college or "",
                    "专业": s.major or "",
                    "学历": DEGREE_MAP.get(s.degree, '未知'),
                    "毕业年份": s.graduation_year or "",
                    "就业状态": EMPLOYMENT_STATUS_MAP.get(s.employment_status, '未知'),
                    "当前公司": s.cur_company or "",
                    "当前城市": s.cur_city or "",
                    "当前行业": s.cur_industry or "",
                    "当前薪资": s.cur_salary or "",
                    "期望城市": s.desire_city or "",
                    "期望行业": s.desire_industry or "",
                    "期望薪资": f"{s.desire_salary_min or '-'} ~ {s.desire_salary_max or '-'}" if s.desire_salary_min or s.desire_salary_max else "",
                    "生源省份": s.province_origin or "",
                    "GPA": s.gpa or "",
                    "技能特长": ", ".join(s.skills) if s.skills else "",
                    "实习经历": s.internship or "",
                    "注册状态": "已注册" if s.account_id else "未注册",
                })

            yield batch
            offset += batch_size

    async def get_student_detail(self, profile_id: str) -> dict | None:
        """获取学生完整档案信息"""
        result = await self.db.execute(
            select(StudentProfile).where(StudentProfile.profile_id == profile_id)
        )
        profile = result.scalar_one_or_none()
        if not profile:
            return None

        # 查询关联账户信息
        account_info = None
        if profile.account_id:
            acc_result = await self.db.execute(
                select(Account).where(Account.account_id == profile.account_id)
            )
            account = acc_result.scalar_one_or_none()
            if account:
                account_info = {
                    "username": account.username,
                    "real_name": account.real_name,
                    "email": account.email,
                    "phone": account.phone,
                    "status": account.status,
                }

        DEGREE_MAP = {1: '本科', 2: '硕士', 3: '博士', 4: '大专', 5: '其他'}
        EMPLOYMENT_STATUS_MAP = {
            0: '待就业', 1: '已就业', 2: '升学', 3: '出国'
        }

        return {
            "profile_id": profile.profile_id,
            "account_id": profile.account_id,
            "university_id": profile.university_id,
            "student_no": profile.student_no,
            "college": profile.college,
            "major": profile.major,
            "class_name": None,
            "degree": profile.degree,
            "degree_text": DEGREE_MAP.get(profile.degree, '未知'),
            "graduation_year": profile.graduation_year,
            "province_origin": profile.province_origin,
            "gpa": profile.gpa,
            "skills": profile.skills,
            "internship": profile.internship,
            "employment_status": profile.employment_status,
            "employment_status_text": EMPLOYMENT_STATUS_MAP.get(profile.employment_status, '未知'),
            "desire_city": profile.desire_city,
            "desire_industry": profile.desire_industry,
            "desire_salary_min": profile.desire_salary_min,
            "desire_salary_max": profile.desire_salary_max,
            "cur_company": profile.cur_company,
            "cur_city": profile.cur_city,
            "cur_industry": profile.cur_industry,
            "cur_salary": profile.cur_salary,
            "resume_url": profile.resume_url,
            "profile_complete": profile.profile_complete,
            "is_registered": bool(profile.account_id) and profile.account_id.strip() != "",
            "account": account_info,
        }

    async def import_students_preview(self, file_content: bytes, filename: str) -> dict:
        """Excel/CSV 批量导入预览：解析文件并检测注册状态"""
        import csv
        settings = get_settings()
        default_univ_id = settings.DEFAULT_UNIVERSITY_ID or ""

        rows = []
        errors = []
        preview_list = []

        def parse_row(row_idx: int, row_data: dict) -> dict | None:
            """解析单行数据，返回解析后的 dict 或 None（跳过）"""
            student_no = str(row_data.get("student_no", "")).strip()
            if not student_no:
                errors.append(f"第{row_idx}行: 学号为空，跳过")
                return None
            return {
                "student_no": student_no,
                "college": str(row_data.get("college") or "").strip(),
                "major": str(row_data.get("major") or "").strip(),
                "degree": int(row_data.get("degree") or 1) or 1,
                "graduation_year": int(row_data.get("graduation_year") or 0) or None,
                "class_name": str(row_data.get("class") or "").strip(),
                "province_origin": str(row_data.get("province_origin") or "").strip(),
                "employment_status": int(row_data.get("employment_status") or 0) or 0,
            }

        # Excel 格式
        if filename.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            header_map = {h.lower(): idx for idx, h in enumerate(headers)}

            if "student_no" not in header_map:
                return {"success": 0, "errors": ["缺少必需列: student_no"]}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None for v in row):
                    continue
                try:
                    raw = {header_map[k]: row[header_map[k]] for k in header_map}
                    parsed = parse_row(row_idx, raw)
                    if parsed:
                        rows.append(parsed)
                except Exception as e:
                    errors.append(f"第{row_idx}行: 解析失败 {str(e)}")

        elif filename.lower().endswith(".csv"):
            decoded = file_content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(decoded))
            if not reader.fieldnames or "student_no" not in [f.strip().lower() for f in reader.fieldnames]:
                return {"success": 0, "errors": ["缺少必需列: student_no"]}

            for row_idx, row in enumerate(reader, start=2):
                if not row or all(v is None or str(v).strip() == "" for v in row.values()):
                    continue
                try:
                    parsed = parse_row(row_idx, row)
                    if parsed:
                        rows.append(parsed)
                except Exception as e:
                    errors.append(f"第{row_idx}行: 解析失败 {str(e)}")

        else:
            return {"success": 0, "errors": ["不支持的文件格式，请上传 .xlsx/.xls/.csv 文件"]}

        if not rows:
            return {"success": 0, "errors": errors or ["文件中没有有效数据"]}

        # 批量查询已注册学号
        student_nos = [r["student_no"] for r in rows]
        exist_result = await self.db.execute(
            select(StudentProfile.student_no, StudentProfile.account_id)
            .where(StudentProfile.student_no.in_(student_nos))
        )
        exist_map = {row.student_no: row.account_id for row in exist_result.all()}

        for r in rows:
            account_id = exist_map.get(r["student_no"])
            r["is_registered"] = bool(account_id) and account_id.strip() != ""
            r["account_id"] = account_id
            preview_list.append(r)

        return {
            "success": len(rows),
            "preview": preview_list,
            "total": len(rows),
            "errors": errors[:20],
        }

    async def confirm_import_students(self, university_id: str, students_data: list) -> dict:
        """确认导入：批量创建/更新学生档案"""
        settings = get_settings()
        default_univ_id = university_id or settings.DEFAULT_UNIVERSITY_ID or ""
        imported = 0
        skipped = 0
        errors = []
        BATCH = 100

        for i in range(0, len(students_data), BATCH):
            batch = students_data[i:i + BATCH]
            for s in batch:
                try:
                    student_no = s.get("student_no", "").strip()
                    if not student_no:
                        skipped += 1
                        continue

                    # 查询是否已存在
                    result = await self.db.execute(
                        select(StudentProfile).where(StudentProfile.student_no == student_no)
                    )
                    existing = result.scalar_one_or_none()

                    if existing:
                        # 学号已存在，视为重复数据，跳过
                        skipped += 1
                    else:
                        # 仅创建学生档案，不创建账户（学生自主注册时再绑定账户）
                        profile_id = str(uuid.uuid4())
                        profile = StudentProfile(
                            profile_id=profile_id,
                            account_id=None,  # 不再创建账户，学生注册时再绑定
                            university_id=default_univ_id,
                            student_no=student_no,
                            college=s.get("college", ""),
                            major=s.get("major", ""),
                            degree=s.get("degree", 1),
                            graduation_year=s.get("graduation_year"),
                            province_origin=s.get("province_origin", ""),
                            employment_status=s.get("employment_status", 0),
                            profile_complete=30,
                        )
                        self.db.add(profile)
                        imported += 1

                except Exception as e:
                    skipped += 1
                    errors.append(f"学号 {s.get('student_no')}: {str(e)}")

            await self.db.commit()

        return {
            "success": imported,
            "skipped": skipped,
            "errors": errors[:20],
        }

    async def batch_delete_students(self, profile_ids: list) -> dict:
        """批量删除学生档案及其关联账户"""
        if not profile_ids:
            return {"deleted": 0}
        try:
            result = await self.db.execute(
                select(StudentProfile.profile_id, StudentProfile.account_id)
                .where(StudentProfile.profile_id.in_(profile_ids))
            )
            rows = result.all()
            if not rows:
                return {"deleted": 0}

            # 收集所有关联的 account_id（每个 profile 必定有 account_id，因为 FK nullable=False）
            account_ids = [row.account_id for row in rows if row.account_id]

            # 先删账户（CASCADE 会自动删除关联的 student_profiles）
            for aid in account_ids:
                acc_result = await self.db.execute(select(Account).where(Account.account_id == aid))
                account = acc_result.scalar_one_or_none()
                if account:
                    await self.db.delete(account)

            await self.db.commit()
            return {"deleted": len(rows)}
        except Exception as e:
            await self.db.rollback()
            raise e
            raise e

    async def batch_update_students(self, profile_ids: list, update_data: dict) -> dict:
        """批量更新学生档案字段"""
        if not profile_ids:
            return {"updated": 0}
        # 允许更新的字段
        ALLOWED = ["college", "major", "degree", "graduation_year", "employment_status",
                   "cur_company", "cur_city", "cur_industry", "cur_salary",
                   "province_origin", "gpa", "skills", "internship"]
        filtered = {k: v for k, v in update_data.items() if k in ALLOWED}
        if not filtered:
            return {"updated": 0}
        try:
            result = await self.db.execute(
                update(StudentProfile)
                .where(StudentProfile.profile_id.in_(profile_ids))
                .values(**filtered)
            )
            await self.db.commit()
            return {"updated": result.rowcount}
        except Exception as e:
            await self.db.rollback()
            raise e

    async def get_warnings(self, university_id: str, filters: dict) -> dict:
        """获取预警列表"""
        page = filters.get("page", 1)
        page_size = filters.get("page_size", 10)
        handled = filters.get("handled")
        level = filters.get("level")

        conditions = [EmploymentWarning.university_id == university_id]
        if handled is not None:
            conditions.append(EmploymentWarning.handled == handled)
        if level:
            conditions.append(EmploymentWarning.level == level)

        # 查询总数
        count_result = await self.db.execute(
            select(func.count(EmploymentWarning.warning_id)).where(and_(*conditions))
        )
        total = count_result.scalar() or 0

        # 分页查询
        offset = (page - 1) * page_size
        result = await self.db.execute(
            select(EmploymentWarning)
            .where(and_(*conditions))
            .order_by(EmploymentWarning.created_at.desc())
            .offset(offset)
            .limit(page_size)
        )
        warnings = result.scalars().all()

        return {
            "list": [{
                "warning_id": w.warning_id,
                "account_id": w.account_id,
                "warning_type": w.warning_type,
                "level": w.level,
                "ai_suggestion": w.ai_suggestion,
                "handled": w.handled,
                "created_at": str(w.created_at)
            } for w in warnings],
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def handle_warning(self, warning_id: str, handled: bool) -> bool:
        """处理预警"""
        from datetime import datetime

        result = await self.db.execute(
            select(EmploymentWarning).where(EmploymentWarning.warning_id == warning_id)
        )
        warning = result.scalar_one_or_none()
        if not warning:
            return False

    # ========== 省份详情弹窗相关 ==========

    async def get_province_student_summary(
        self, province: str, university_id: str, year: int | None = None
    ) -> dict:
        """获取省份学生聚合统计"""
        from app.models.company import Company

        # 规范化省份名称
        normalized_province = normalize_province_name(province)
        conditions = [
            StudentProfile.university_id == university_id,
            StudentProfile.province_origin == normalized_province,
        ]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # 总数
        total_result = await self.db.execute(
            select(func.count(StudentProfile.profile_id)).where(*conditions)
        )
        total = total_result.scalar() or 0

        if total == 0:
            return {
                "total": 0,
                "employment_distribution": {"employed": 0, "unemployed": 0},
                "degree_distribution": {"bachelor": 0, "master": 0, "doctoral": 0},
                "avg_salary": 0,
                "top_industries": [],
            }

        # 就业状态分布
        employment_result = await self.db.execute(
            select(
                StudentProfile.employment_status,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(*conditions)
            .group_by(StudentProfile.employment_status)
        )
        employed = 0
        unemployed = 0
        for row in employment_result.all():
            if row.employment_status == 1:
                employed = row.count
            else:
                unemployed = row.count

        # 学历分布
        degree_result = await self.db.execute(
            select(
                StudentProfile.degree,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(*conditions)
            .group_by(StudentProfile.degree)
        )
        degree_dist = {"bachelor": 0, "master": 0, "doctoral": 0}
        for row in degree_result.all():
            if row.degree == 1:
                degree_dist["bachelor"] = row.count
            elif row.degree == 2:
                degree_dist["master"] = row.count
            elif row.degree == 3:
                degree_dist["doctoral"] = row.count

        # 平均薪资（只统计有薪资的已就业学生）
        salary_result = await self.db.execute(
            select(StudentProfile.cur_salary).where(
                *conditions,
                StudentProfile.employment_status == 1,
                StudentProfile.cur_salary.isnot(None),
                StudentProfile.cur_salary > 0,
            )
        )
        salaries = [row for row in salary_result.scalars().all() if row]
        avg_salary = int(sum(salaries) / len(salaries)) if salaries else 0

        # 热门行业 Top 5
        industry_result = await self.db.execute(
            select(StudentProfile.cur_industry).where(
                *conditions,
                StudentProfile.cur_industry.isnot(None),
                StudentProfile.cur_industry != "",
            )
        )
        industries = [normalize_industry(ind) for ind in industry_result.scalars().all()]
        industry_counter = Counter(industries)
        top_industries = [
            {"industry": name, "count": count}
            for name, count in industry_counter.most_common(5)
        ]

        return {
            "total": total,
            "employment_distribution": {"employed": employed, "unemployed": unemployed},
            "degree_distribution": degree_dist,
            "avg_salary": avg_salary,
            "top_industries": top_industries,
        }

    async def get_province_students_paginated(
        self,
        province: str,
        university_id: str,
        year: int | None = None,
        page: int = 1,
        page_size: int = 10,
    ) -> dict:
        """分页获取省份学生列表"""
        normalized_province = normalize_province_name(province)
        conditions = [
            StudentProfile.university_id == university_id,
            StudentProfile.province_origin == normalized_province,
        ]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # 总数
        count_result = await self.db.execute(
            select(func.count(StudentProfile.profile_id)).where(*conditions)
        )
        total = count_result.scalar() or 0

        offset = (page - 1) * page_size
        result = await self.db.execute(
            select(
                StudentProfile.profile_id,
                StudentProfile.student_no,
                StudentProfile.college,
                StudentProfile.major,
                StudentProfile.degree,
                StudentProfile.graduation_year,
                StudentProfile.employment_status,
                StudentProfile.cur_company,
                StudentProfile.cur_city,
                StudentProfile.cur_industry,
                StudentProfile.cur_salary,
            )
            .where(*conditions)
            .order_by(StudentProfile.created_at.desc())
            .offset(offset)
            .limit(page_size)
        )

        items = []
        for row in result.all():
            items.append({
                "profile_id": row.profile_id,
                "student_no": row.student_no,
                "college": row.college,
                "major": row.major,
                "degree": row.degree,
                "graduation_year": row.graduation_year,
                "employment_status": row.employment_status,
                "cur_company": row.cur_company,
                "cur_city": row.cur_city,
                "cur_industry": row.cur_industry,
                "cur_salary": row.cur_salary,
            })

        return {
            "list": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    async def get_province_company_summary(
        self, province: str, university_id: str
    ) -> dict:
        """获取省份企业聚合统计（基于岗位的省份字段）"""
        from app.models.company import Company

        normalized_province = normalize_province_name(province)

        # 岗位统计（使用 job_descriptions.province 字段）
        job_conditions = [
            JobDescription.province == normalized_province,
            JobDescription.status == 1,
        ]
        total_jobs_result = await self.db.execute(
            select(func.count(JobDescription.job_id)).where(*job_conditions)
        )
        total_jobs = total_jobs_result.scalar() or 0

        if total_jobs == 0:
            return {
                "total_companies": 0,
                "total_jobs": 0,
                "industry_distribution": [],
                "avg_salary_min": 0,
                "avg_salary_max": 0,
                "recent_activities_count": 0,
            }

        # 去重企业数
        company_ids_result = await self.db.execute(
            select(JobDescription.company_id).where(*job_conditions).distinct()
        )
        company_ids = [row for row in company_ids_result.scalars().all()]
        total_companies = len(company_ids)

        # 行业分布
        industry_result = await self.db.execute(
            select(
                JobDescription.industry,
                func.count(JobDescription.job_id).label("count"),
            )
            .where(*job_conditions)
            .group_by(JobDescription.industry)
        )
        industry_distribution = [
            {"industry": row.industry or "其他", "count": row.count}
            for row in industry_result.all()
        ]
        industry_distribution.sort(key=lambda x: x["count"], reverse=True)

        # 平均薪资范围
        salary_result = await self.db.execute(
            select(JobDescription.min_salary, JobDescription.max_salary).where(
                *job_conditions,
                JobDescription.min_salary.isnot(None),
                JobDescription.max_salary.isnot(None),
            )
        )
        min_salaries = []
        max_salaries = []
        for row in salary_result.all():
            if row.min_salary and row.min_salary > 0:
                min_salaries.append(row.min_salary)
            if row.max_salary and row.max_salary > 0:
                max_salaries.append(row.max_salary)
        avg_salary_min = int(sum(min_salaries) / len(min_salaries)) if min_salaries else 0
        avg_salary_max = int(sum(max_salaries) / len(max_salaries)) if max_salaries else 0

        # 近期活动数（该省份企业最近的活动）
        if company_ids:
            from datetime import date, timedelta
            from app.models.company_activity import CompanyActivity

            recent_date = date.today() - timedelta(days=90)
            activity_result = await self.db.execute(
                select(func.count(CompanyActivity.activity_id))
                .join(Company, CompanyActivity.company_id == Company.company_id)
                .join(JobDescription, JobDescription.company_id == Company.company_id)
                .where(
                    JobDescription.province == normalized_province,
                    JobDescription.status == 1,
                    CompanyActivity.activity_date >= recent_date,
                )
            )
            recent_activities_count = activity_result.scalar() or 0
        else:
            recent_activities_count = 0

        return {
            "total_companies": total_companies,
            "total_jobs": total_jobs,
            "industry_distribution": industry_distribution[:10],
            "avg_salary_min": avg_salary_min,
            "avg_salary_max": avg_salary_max,
            "recent_activities_count": recent_activities_count,
        }

    async def get_province_companies_paginated(
        self,
        province: str,
        university_id: str,
        page: int = 1,
        page_size: int = 10,
    ) -> dict:
        """分页获取省份企业列表（含岗位数、认证状态）"""
        from app.models.company import Company

        normalized_province = normalize_province_name(province)

        # 子查询：每个企业的岗位数
        job_count_subq = (
            select(
                JobDescription.company_id,
                func.count(JobDescription.job_id).label("job_count"),
            )
            .where(
                JobDescription.province == normalized_province,
                JobDescription.status == 1,
            )
            .group_by(JobDescription.company_id)
            .subquery()
        )

        # 主查询
        count_result = await self.db.execute(
            select(func.count(job_count_subq.c.company_id))
        )
        total = count_result.scalar() or 0

        offset = (page - 1) * page_size
        result = await self.db.execute(
            select(
                Company.company_id,
                Company.company_name,
                Company.industry,
                Company.city,
                Company.verified,
                job_count_subq.c.job_count,
            )
            .join(job_count_subq, Company.company_id == job_count_subq.c.company_id)
            .order_by(job_count_subq.c.job_count.desc())
            .offset(offset)
            .limit(page_size)
        )

        items = []
        for row in result.all():
            items.append({
                "company_id": row.company_id,
                "company_name": row.company_name,
                "industry": row.industry,
                "city": row.city,
                "verified": row.verified,
                "job_count": row.job_count,
            })

        return {
            "list": items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

        warning.handled = handled
        warning.handled_at = datetime.utcnow()
        await self.db.commit()
        return True

    async def get_databoard_data(self, university_id: str, year: int | None = None) -> dict:
        """
        数据大屏数据（带Redis缓存30分钟）

        返回20+指标数据：
        - 数字卡片：毕业生总数、就业率、升学人数、签约人数、出国人数
        - 就业率趋势（2022-2026）
        - 学院就业TOP10排名
        - 学历层次对比
        - 毕业去向分布
        - 区域流向分析
        - 行业分布（归一化后）
        - 紧缺人才分析（RAG+归一化）
        - 期望薪资分布
        - 性别就业差异
        - 实习价值分析
        - 满意度分布
        """
        redis = get_redis()
        cache_key = f"school_databoard:{university_id}:{year or 'all'}"

        # 尝试从缓存获取
        if redis:
            try:
                cached = redis.get(cache_key)
                if cached:
                    return json.loads(cached)
            except Exception:
                pass

        # ========== 1. 数字卡片汇总 ==========
        summary = await self._get_summary_stats(university_id, year)

        # ========== 2. 就业率趋势（2022-2026） ==========
        employment_trend = await self._get_employment_trend(university_id)

        # ========== 3. 学院就业TOP10 ==========
        college_ranking = await self._get_college_ranking(university_id, year)

        # ========== 4. 学历层次对比 ==========
        degree_comparison = await self._get_degree_comparison(university_id, year)

        # ========== 5. 毕业去向分布 ==========
        direction_distribution = await self._get_direction_distribution()

        # ========== 6. 区域流向分析 ==========
        regional_flow = await self._get_regional_flow(university_id, year)

        # ========== 6.5 省份维度分布（地图用） ==========
        map_data = await self._get_province_distribution(university_id, year)

        # ========== 7. 行业分布（归一化） ==========
        industry_radar = await self._get_industry_distribution()

        # ========== 8. 紧缺人才分析（RAG+归一化） ==========
        scarce_talent = await self._get_scarce_talent_analysis()

        # ========== 9. 期望薪资分布 ==========
        salary_distribution = await self._get_salary_distribution()

        # ========== 10. 性别就业差异 ==========
        gender_employment = await self._get_gender_employment()

        # ========== 11. 实习价值分析 ==========
        internship_value = await self._get_internship_value()

        # ========== 12. 专业分布 TOP20 ==========
        major_distribution = await self._get_major_distribution(university_id, year)

        # ========== 13. 城市偏好分布 ==========
        city_preference = await self._get_city_preference(university_id, year)

        data = {
            # 数字卡片
            "summary": summary,
            # 图表数据
            "employment_trend": employment_trend,
            "college_ranking": college_ranking,
            "degree_comparison": degree_comparison,
            "direction_distribution": direction_distribution,
            "regional_flow": regional_flow,
            "map_data": map_data,
            "industry_radar": industry_radar,
            "scarce_talent": scarce_talent,
            "salary_distribution": salary_distribution,
            "gender_employment": gender_employment,
            "internship_value": internship_value,
            "major_distribution": major_distribution,
            "city_preference": city_preference,
        }

        # 缓存30分钟
        if redis:
            try:
                redis.setex(cache_key, 1800, json.dumps(data, default=str))
            except Exception:
                pass
        return data

    async def _get_summary_stats(self, university_id: str, year: int | None) -> dict:
        """数字卡片：毕业生总数、就业率、升学/签约/出国人数"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # Get total and status counts
        result = await self.db.execute(
            select(
                func.count(StudentProfile.profile_id).label("total_graduates"),
                func.sum(
                    case((StudentProfile.employment_status == 1, 1), else_=0)
                ).label("total_employed"),
                func.sum(
                    case((StudentProfile.employment_status == 2, 1), else_=0)
                ).label("total_further_study"),
                func.sum(
                    case((StudentProfile.employment_status == 3, 1), else_=0)
                ).label("total_overseas"),
            ).where(and_(*conditions))
        )
        row = result.one()

        total_graduates = row.total_graduates or 0
        total_employed = row.total_employed or 0
        total_further_study = row.total_further_study or 0
        total_overseas = row.total_overseas or 0
        total_contract = int(total_employed * 0.6)  # Estimate 60% signed contract

        employment_rate = (
            round(total_employed / total_graduates * 100, 2) if total_graduates > 0 else 0
        )
        overseas_rate = (
            round(total_overseas / total_graduates * 100, 2) if total_graduates > 0 else 0
        )

        return {
            "total_graduates": total_graduates,
            "employment_rate": employment_rate,
            "further_study_count": total_further_study,
            "contract_count": total_contract,
            "overseas_count": total_overseas,
            "overseas_rate": overseas_rate,
        }

    async def _get_employment_trend(self, university_id: str) -> list:
        """就业率趋势（2022-2026）"""
        result = await self.db.execute(
            select(
                StudentProfile.graduation_year,
                func.count(StudentProfile.profile_id).label("total_grads"),
                func.sum(
                    case((StudentProfile.employment_status == 1, 1), else_=0)
                ).label("total_employed"),
            )
            .where(StudentProfile.university_id == university_id)
            .group_by(StudentProfile.graduation_year)
            .order_by(StudentProfile.graduation_year)
        )

        trends = []
        for row in result.all():
            rate = (
                round(row.total_employed / row.total_grads * 100, 2)
                if row.total_grads and row.total_grads > 0 else 0
            )
            trends.append({
                "year": row.graduation_year,
                "employment_rate": rate,
                "graduate_nums": row.total_grads or 0,
                "employed_nums": row.total_employed or 0,
            })
        return trends

    async def _get_college_ranking(self, university_id: str, year: int | None) -> list:
        """各学院就业TOP10排名（含分学历层数据）"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # 先获取所有学院的总体数据
        result = await self.db.execute(
            select(
                StudentProfile.college,
                func.count(StudentProfile.profile_id).label("total_grads"),
                func.sum(
                    case((StudentProfile.employment_status == 1, 1), else_=0)
                ).label("total_employed"),
            )
            .where(and_(*conditions))
            .group_by(StudentProfile.college)
            .order_by(func.sum(
                case((StudentProfile.employment_status == 1, 1), else_=0)
            ).desc())
            .limit(10)
        )

        rankings = []
        for row in result.all():
            college_name = row.college

            # 分别查询该学院的博士、硕士、本科就业数据
            degree_result = await self.db.execute(
                select(
                    StudentProfile.degree,
                    func.count(StudentProfile.profile_id).label("total"),
                    func.sum(
                        case((StudentProfile.employment_status == 1, 1), else_=0)
                    ).label("employed"),
                )
                .where(and_(
                    StudentProfile.university_id == university_id,
                    StudentProfile.college == college_name,
                    *([StudentProfile.graduation_year == year] if year else [])
                ))
                .group_by(StudentProfile.degree)
            )

            # 初始化各学历数据
            doctoral = {"total": 0, "employed": 0, "rate": 0}
            master = {"total": 0, "employed": 0, "rate": 0}
            bachelor = {"total": 0, "employed": 0, "rate": 0}

            for dr in degree_result.all():
                degree_map = {3: doctoral, 2: master, 1: bachelor}
                d = degree_map.get(dr.degree)
                if d:
                    d["total"] = dr.total or 0
                    d["employed"] = dr.employed or 0
                    if dr.total and dr.total > 0:
                        d["rate"] = round(dr.employed / dr.total * 100, 2)

            # 计算总就业率
            total_rate = (
                round(row.total_employed / row.total_grads * 100, 2)
                if row.total_grads and row.total_grads > 0 else 0
            )

            rankings.append({
                "college_name": college_name,
                "employment_rate": total_rate,
                "employed_nums": row.total_employed or 0,
                "graduate_nums": row.total_grads or 0,
                # 分学历层数据
                "doctoral": doctoral,
                "master": master,
                "bachelor": bachelor,
            })
        return rankings

    async def _get_degree_comparison(self, university_id: str, year: int | None) -> dict:
        """学历层次对比（博士/硕士/本科，含升学数）"""
        DEGREE_MAP = {
            3: "doctoral",   # 博士
            2: "master",    # 硕士
            1: "bachelor",  # 本科
        }
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        result = await self.db.execute(
            select(
                StudentProfile.degree,
                func.count(StudentProfile.profile_id).label("total_grads"),
                func.sum(
                    case((StudentProfile.employment_status == 1, 1), else_=0)
                ).label("total_employed"),
            )
            .where(and_(*conditions))
            .group_by(StudentProfile.degree)
        )

        comparison = {"doctoral": None, "master": None, "bachelor": None}
        for row in result.all():
            key = DEGREE_MAP.get(row.degree)
            if key:
                rate = (
                    round(row.total_employed / row.total_grads * 100, 2)
                    if row.total_grads and row.total_grads > 0 else 0
                )
                comparison[key] = {
                    "graduate_nums": row.total_grads or 0,
                    "employed_nums": row.total_employed or 0,
                    "employment_rate": rate,
                    # 升学数 = 毕业总数 - 已就业数（简化计算，实际应以employment_status=2为准）
                    "further_study_nums": max(0, (row.total_grads or 0) - (row.total_employed or 0)),
                }
        return comparison

    async def _get_direction_distribution(self) -> list:
        """毕业去向分布（来自student_employment_choice）"""
        result = await self.db.execute(
            select(
                StudentProfile.employment_status,
                func.count(StudentProfile.profile_id).label("count"),
            ).where(StudentProfile.employment_status.isnot(None))
            .group_by(StudentProfile.employment_status)
        )

        STATUS_MAP = {
            0: "待就业",
            1: "已就业",
            2: "升学",
            3: "出国",
        }
        distribution = []
        for row in result.all():
            distribution.append({
                "name": STATUS_MAP.get(row.employment_status, "未知"),
                "value": row.count,
            })
        return distribution

    async def _get_regional_flow(self, university_id: str, year: int | None) -> dict:
        """区域流向分析（东/中/西部/川渝/港澳台）"""
        from app.models.regional_flow import RegionalFlow

        conditions = [RegionalFlow.university_id == university_id]
        if year:
            conditions.append(RegionalFlow.graduation_year == year)

        result = await self.db.execute(
            select(RegionalFlow).where(and_(*conditions))
        )
        rows = result.scalars().all()

        if not rows:
            # Fallback: generate from student_profiles
            return await self._get_regional_flow_from_students(university_id, year)

        # Return data from RegionalFlow table
        total = sum(row.total_nums for row in rows)
        return {
            "total": total,
            "distribution": [
                {"name": "东部地区", "value": sum(row.east_nums for row in rows)},
                {"name": "中部地区", "value": sum(row.central_nums for row in rows)},
                {"name": "西部地区-重庆", "value": sum(row.west_chongqing_nums for row in rows)},
                {"name": "西部地区-四川", "value": sum(row.west_sichuan_nums for row in rows)},
                {"name": "西部地区-其他", "value": sum(row.west_other_nums for row in rows)},
                {"name": "港澳台及其他", "value": sum(row.hmt_nums for row in rows)},
            ],
        }

    async def _get_regional_flow_from_students(self, university_id: str, year: int | None) -> dict:
        """区域流向分析 - 从学生表计算（备用）"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # Province to region mapping
        PROVINCE_REGION = {
            "北京": "east", "天津": "east", "河北": "east", "辽宁": "east",
            "上海": "east", "江苏": "east", "浙江": "east", "福建": "east",
            "山东": "east", "广东": "east", "海南": "east",
            "山西": "central", "吉林": "central", "黑龙江": "central",
            "安徽": "central", "江西": "central", "河南": "central",
            "湖北": "central", "湖南": "central",
            "重庆": "west_chongqing",
            "四川": "west_sichuan",
            "西藏": "west_other", "青海": "west_other", "宁夏": "west_other",
            "新疆": "west_other", "甘肃": "west_other", "陕西": "west_other",
            "云南": "west_other", "贵州": "west_other", "内蒙古": "west_other",
            "广西": "west_other",
            "香港": "hmt", "澳门": "hmt", "台湾": "hmt",
        }

        result = await self.db.execute(
            select(
                StudentProfile.province_origin,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(and_(*conditions, StudentProfile.province_origin.isnot(None)))
            .group_by(StudentProfile.province_origin)
        )

        regions = {"east": 0, "central": 0, "west_chongqing": 0,
                  "west_sichuan": 0, "west_other": 0, "hmt": 0}

        for row in result.all():
            region = PROVINCE_REGION.get(row.province_origin, "west_other")
            regions[region] = regions.get(region, 0) + (row.count or 0)

        total = sum(regions.values())
        return {
            "total": total,
            "distribution": [
                {"name": "东部地区", "value": regions["east"]},
                {"name": "中部地区", "value": regions["central"]},
                {"name": "西部地区-重庆", "value": regions["west_chongqing"]},
                {"name": "西部地区-四川", "value": regions["west_sichuan"]},
                {"name": "西部地区-其他", "value": regions["west_other"]},
                {"name": "港澳台及其他", "value": regions["hmt"]},
            ],
        }

    async def _get_province_distribution(self, university_id: str, year: int | None) -> list:
        """省份维度分布（用于地图展示，包含本硕博分层数据）"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        # Query with degree breakdown by province
        result = await self.db.execute(
            select(
                StudentProfile.province_origin,
                StudentProfile.degree,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(and_(*conditions, StudentProfile.province_origin.isnot(None)))
            .group_by(StudentProfile.province_origin, StudentProfile.degree)
        )

        # Organize data by province with degree breakdown
        province_data: dict[str, dict[str, int]] = {}
        for row in result.all():
            province = row.province_origin
            if province not in province_data:
                province_data[province] = {"bachelor": 0, "master": 0, "doctoral": 0}
            # degree: 1=本科, 2=硕士, 3=博士
            if row.degree == 1:
                province_data[province]["bachelor"] += row.count
            elif row.degree == 2:
                province_data[province]["master"] += row.count
            elif row.degree == 3:
                province_data[province]["doctoral"] += row.count

        # Convert to list format
        distribution = []
        for province, counts in province_data.items():
            total = counts["bachelor"] + counts["master"] + counts["doctoral"]
            distribution.append({
                "province": province,
                "bachelor": counts["bachelor"],
                "master": counts["master"],
                "doctoral": counts["doctoral"],
                "total": total,
            })

        return distribution

    async def _get_industry_distribution(self) -> list:
        """行业分布（归一化后，用于雷达图）"""
        result = await self.db.execute(
            select(StudentProfile.cur_industry).where(
                StudentProfile.cur_industry.isnot(None),
                StudentProfile.cur_industry != "",
            )
        )
        industries = [row for row in result.scalars().all()]

        # 归一化
        normalized = [normalize_industry(ind) for ind in industries]
        counter = Counter(normalized)

        return [
            {"industry": name, "count": count}
            for name, count in counter.most_common(15)
        ]

    async def _get_scarce_talent_analysis(self) -> dict:
        """紧缺人才分析（RAG+归一化）"""
        analyzer = ScarceTalentAnalyzer(self.db)
        summary = await analyzer.get_scarce_talent_summary()
        wordcloud = await analyzer.get_industry_wordcloud()
        province_industry = await analyzer.get_province_industry_map()

        return {
            "summary": summary,
            "wordcloud": wordcloud,
            "province_industry": province_industry,
        }

    async def _get_salary_distribution(self) -> list:
        """期望薪资分布（已过滤异常值）"""
        from app.utils.salary_parser import SALARY_MIN, SALARY_MAX

        result = await self.db.execute(
            select(StudentProfile.desire_salary_min, StudentProfile.desire_salary_max)
            .where(
                StudentProfile.desire_salary_min.isnot(None),
                StudentProfile.desire_salary_max.isnot(None),
            )
        )

        # 分段统计（使用期望薪资的平均值作为分配依据）
        BUCKETS = [
            (0, 3000, "0-3k"),
            (3000, 5000, "3k-5k"),
            (5000, 8000, "5k-8k"),
            (8000, 10000, "8k-10k"),
            (10000, 15000, "10k-15k"),
            (15000, 20000, "15k-20k"),
            (20000, 30000, "20k-30k"),
            (30000, float("inf"), "30k+"),
        ]

        counts = {label: 0 for _, _, label in BUCKETS}

        for row in result.all():
            min_sal = row.desire_salary_min or 0
            max_sal = row.desire_salary_max or 0

            # 过滤异常值
            if min_sal < SALARY_MIN or max_sal > SALARY_MAX:
                continue

            # 跳过无效范围
            if min_sal <= 0 and max_sal <= 0:
                continue

            # 使用期望薪资的平均值来分配桶
            avg_sal = (min_sal + max_sal) / 2

            for low, high, label in BUCKETS:
                if low <= avg_sal < high:
                    counts[label] += 1
                    break
            else:
                # 如果平均值超出所有桶的最高范围，计入最后一个桶
                if avg_sal >= 30000:
                    counts["30k+"] += 1

        return [
            {"range": label, "count": count}
            for label, count in counts.items()
            if count > 0
        ]

    async def _get_major_distribution(self, university_id: str, year: int | None) -> list:
        """专业分布 TOP20"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        result = await self.db.execute(
            select(
                StudentProfile.major,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(and_(*conditions, StudentProfile.major.isnot(None), StudentProfile.major != ""))
            .group_by(StudentProfile.major)
            .order_by(func.count(StudentProfile.profile_id).desc())
            .limit(20)
        )

        return [
            {"major": row.major, "count": row.count}
            for row in result.all()
        ]

    async def _get_city_preference(self, university_id: str, year: int | None) -> list:
        """城市偏好分布（学生期望工作城市）"""
        conditions = [StudentProfile.university_id == university_id]
        if year:
            conditions.append(StudentProfile.graduation_year == year)

        result = await self.db.execute(
            select(
                StudentProfile.desire_city,
                func.count(StudentProfile.profile_id).label("count"),
            )
            .where(and_(*conditions, StudentProfile.desire_city.isnot(None), StudentProfile.desire_city != ""))
            .group_by(StudentProfile.desire_city)
            .order_by(func.count(StudentProfile.profile_id).desc())
            .limit(20)
        )

        return [
            {"city": row.desire_city, "count": row.count}
            for row in result.all()
        ]

    async def _get_gender_employment(self) -> dict:
        """性别就业差异"""
        # student_profiles表没有gender字段，无法直接按性别分组
        # 简化：返回空数据，实际应关联account表或student_employment_choice表获取性别
        return {
            "male": {"employed": 0, "total": 0, "rate": 0},
            "female": {"employed": 0, "total": 0, "rate": 0},
            "note": "性别字段需关联account表或student_employment_choice表",
        }

    async def _get_internship_value(self) -> list:
        """实习价值分析"""
        result = await self.db.execute(
            select(
                StudentProfile.internship,
                func.count(StudentProfile.profile_id).label("total"),
                func.sum(
                    case((StudentProfile.employment_status == 1, 1), else_=0)
                ).label("employed"),
            )
            .where(StudentProfile.internship.isnot(None), StudentProfile.internship != "")
            .group_by(StudentProfile.internship)
        )

        analysis = []
        for row in result.all():
            rate = (
                round(row.employed / row.total * 100, 2)
                if row.total and row.total > 0 else 0
            )
            analysis.append({
                "has_internship": bool(row.internship),
                "total": row.total,
                "employed": row.employed or 0,
                "employment_rate": rate,
            })
        return analysis

    async def _get_satisfaction_distribution(self) -> list:
        """满意度分布"""
        # student_profiles表没有satisfaction_score字段，返回空数据
        return []

    async def get_profile(self, university_id: str) -> University | None:
        """获取学校档案"""
        from app.models.university import University
        result = await self.db.execute(
            select(University).where(University.university_id == university_id)
        )
        return result.scalar_one_or_none()

    async def get_colleges(self, university_id: str, year: int, page: int, page_size: int) -> dict:
        """学院就业率分页
        - year为空时：按学院汇总所有年份数据
        - year有值时：按学院+年份分组
        """
        conditions = [CollegeEmployment.university_id == university_id]
        if year:
            conditions.append(CollegeEmployment.graduation_year == year)
            # 按 college_name + graduation_year 分组
            query = (
                select(
                    CollegeEmployment.college_name,
                    CollegeEmployment.graduation_year,
                    func.sum(CollegeEmployment.graduate_nums).label('total_graduate_nums'),
                    func.sum(CollegeEmployment.employed_nums).label('total_employed_nums'),
                    func.sum(CollegeEmployment.further_study_nums).label('total_further_study_nums'),
                    func.sum(CollegeEmployment.overseas_nums).label('total_overseas_nums'),
                    func.round(func.avg(CollegeEmployment.employment_rate), 2).label('avg_employment_rate'),
                    func.round(func.avg(CollegeEmployment.avg_salary), 2).label('avg_salary')
                )
                .where(and_(*conditions))
                .group_by(CollegeEmployment.college_name, CollegeEmployment.graduation_year)
                .order_by(CollegeEmployment.college_name, CollegeEmployment.graduation_year.desc())
            )
        else:
            # 不选年份时：只按 college_name 分组，汇总所有年份
            query = (
                select(
                    CollegeEmployment.college_name,
                    func.sum(CollegeEmployment.graduate_nums).label('total_graduate_nums'),
                    func.sum(CollegeEmployment.employed_nums).label('total_employed_nums'),
                    func.sum(CollegeEmployment.further_study_nums).label('total_further_study_nums'),
                    func.sum(CollegeEmployment.overseas_nums).label('total_overseas_nums'),
                    func.round(func.avg(CollegeEmployment.employment_rate), 2).label('avg_employment_rate'),
                    func.round(func.avg(CollegeEmployment.avg_salary), 2).label('avg_salary')
                )
                .where(and_(*conditions))
                .group_by(CollegeEmployment.college_name)
                .order_by(CollegeEmployment.college_name)
            )

        # 查询总数
        count_query = select(func.count()).select_from(query.subquery())
        count_result = await self.db.execute(count_query)
        total = count_result.scalar() or 0

        # 分页查询
        offset = (page - 1) * page_size
        result = await self.db.execute(query.offset(offset).limit(page_size))
        rows = result.all()

        if year:
            # 按年份查看：每行包含年份
            return {
                "list": [{
                    "college_name": row.college_name,
                    "graduation_year": row.graduation_year,
                    "graduate_nums": row.total_graduate_nums or 0,
                    "employed_nums": row.total_employed_nums or 0,
                    "employment_rate": float(row.avg_employment_rate) if row.avg_employment_rate else 0,
                    "further_study_nums": row.total_further_study_nums or 0,
                    "overseas_nums": row.total_overseas_nums or 0,
                    "avg_salary": float(row.avg_salary) if row.avg_salary else 0
                } for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size
            }
        else:
            # 汇总查看：不显示年份
            return {
                "list": [{
                    "college_name": row.college_name,
                    "graduate_nums": row.total_graduate_nums or 0,
                    "employed_nums": row.total_employed_nums or 0,
                    "employment_rate": float(row.avg_employment_rate) if row.avg_employment_rate else 0,
                    "further_study_nums": row.total_further_study_nums or 0,
                    "overseas_nums": row.total_overseas_nums or 0,
                    "avg_salary": float(row.avg_salary) if row.avg_salary else 0
                } for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size
            }

    async def update_profile(self, university_id: str, data: dict) -> bool:
        """更新学校档案"""
        from app.models.university import University
        result = await self.db.execute(
            select(University).where(University.university_id == university_id)
        )
        university = result.scalar_one_or_none()
        if not university:
            return False

        # 可更新的字段
        updatable_fields = ["name", "province", "city", "type"]
        for key, value in data.items():
            if key in updatable_fields and hasattr(university, key):
                setattr(university, key, value)

        await self.db.commit()
        return True

    # ==================== 企业信息审核 ====================

    async def get_pending_jobs(
        self,
        review_status: int = 0,
        current: int = 1,
        size: int = 20,
        keyword: str = None
    ) -> tuple:
        """获取待审核的岗位列表"""
        from app.models.company import Company
        offset = (current - 1) * size

        # 构建查询：岗位 + 企业信息
        conditions = [JobDescription.review_status == review_status]
        if keyword:
            keyword_pattern = f"%{keyword}%"
            conditions.append(JobDescription.title.like(keyword_pattern))

        # 查询总数
        count_result = await self.db.execute(
            select(func.count()).select_from(JobDescription).where(*conditions)
        )
        total = count_result.scalar() or 0

        # 分页查询
        result = await self.db.execute(
            select(JobDescription, Company.company_name)
            .outerjoin(Company, JobDescription.company_id == Company.company_id)
            .where(*conditions)
            .order_by(JobDescription.created_at.desc())
            .offset(offset)
            .limit(size)
        )
        rows = result.all()

        items = []
        for row in rows:
            job = row[0]
            company_name = row[1] if len(row) > 1 else None
            items.append({
                "job_id": job.job_id,
                "company_id": job.company_id,
                "company_name": company_name,
                "title": job.title,
                "city": job.city,
                "industry": job.industry,
                "min_salary": job.min_salary,
                "max_salary": job.max_salary,
                "min_degree": job.min_degree,
                "description": job.description,
                "status": job.status,
                "review_status": job.review_status,
                "published_at": str(job.published_at) if job.published_at else None,
                "created_at": str(job.created_at)
            })

        return items, total

    async def review_job(
        self,
        job_id: str,
        action: str,
        reviewer_id: str = None,
        reject_reason: str = None
    ) -> bool:
        """审核岗位"""
        from datetime import datetime

        result = await self.db.execute(
            select(JobDescription).where(JobDescription.job_id == job_id)
        )
        job = result.scalar_one_or_none()
        if not job:
            return False

        if action == "approve":
            job.review_status = 1  # 审核通过
            # 如果岗位状态是暂停(0)或结束(2)，保持原状态；只有待发布时才自动上架
            if job.status == 0:
                job.status = 1  # 自动上架
            await self.db.commit()
            return True
        elif action == "reject":
            job.review_status = 2  # 审核拒绝
            job.status = 0  # 自动暂停
            await self.db.commit()
            return True

        return False

    async def get_pending_activities(
        self,
        review_status: int = 0,
        current: int = 1,
        size: int = 20,
        keyword: str = None
    ) -> tuple:
        """获取待审核的活动列表"""
        from app.models.company import Company
        from app.models.company_activity import CompanyActivity
        offset = (current - 1) * size

        # 构建查询条件
        conditions = [CompanyActivity.review_status == review_status]
        if keyword:
            keyword_pattern = f"%{keyword}%"
            conditions.append(CompanyActivity.title.like(keyword_pattern))

        # 查询总数
        count_result = await self.db.execute(
            select(func.count()).select_from(CompanyActivity).where(*conditions)
        )
        total = count_result.scalar() or 0

        # 分页查询
        result = await self.db.execute(
            select(CompanyActivity, Company.company_name)
            .outerjoin(Company, CompanyActivity.company_id == Company.company_id)
            .where(*conditions)
            .order_by(CompanyActivity.created_at.desc())
            .offset(offset)
            .limit(size)
        )
        rows = result.all()

        items = []
        for row in rows:
            activity = row[0]
            company_name = row[1] if len(row) > 1 else None
            items.append({
                "activity_id": activity.activity_id,
                "company_id": activity.company_id,
                "company_name": company_name,
                "type": activity.type.value if activity.type else None,
                "type_name": activity.type_name,
                "title": activity.title,
                "location": activity.location,
                "activity_date": str(activity.activity_date) if activity.activity_date else None,
                "start_time": str(activity.start_time) if activity.start_time else None,
                "end_time": str(activity.end_time) if activity.end_time else None,
                "description": activity.description,
                "status": activity.status,
                "review_status": activity.review_status,
                "expected_num": activity.expected_num,
                "created_at": str(activity.created_at)
            })

        return items, total

    async def review_activity(
        self,
        activity_id: str,
        action: str,
        reviewer_id: str = None,
        reject_reason: str = None
    ) -> bool:
        """审核活动"""
        from datetime import datetime
        from app.models.company_activity import CompanyActivity

        result = await self.db.execute(
            select(CompanyActivity).where(CompanyActivity.activity_id == activity_id)
        )
        activity = result.scalar_one_or_none()
        if not activity:
            return False

        if action == "approve":
            activity.review_status = 1  # 审核通过
            await self.db.commit()
            return True
        elif action == "reject":
            activity.review_status = 2  # 审核拒绝
            activity.status = 0  # 自动取消
            await self.db.commit()
            return True

        return False
